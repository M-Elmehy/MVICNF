import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill
from io import BytesIO
from pathlib import Path


# === Excel Generation Logic ===
def generate_excel(devices, blocks, rows_per_block, node_seq, rules):
    param_template = [
        "MCM.CONFIG.Port1MasterCmd[{i}].Enable",
        "MCM.CONFIG.Port1MasterCmd[{i}].IntAddress",
        "MCM.CONFIG.Port1MasterCmd[{i}].PollInt",
        "MCM.CONFIG.Port1MasterCmd[{i}].Count",
        "MCM.CONFIG.Port1MasterCmd[{i}].Swap",
        "MCM.CONFIG.Port1MasterCmd[{i}].Node",
        "MCM.CONFIG.Port1MasterCmd[{i}].Func",
        "MCM.CONFIG.Port1MasterCmd[{i}].DevAddress",
    ]

    count_map = rules["count_map"]
    devaddr_map = rules["devaddr_map"]
    func_value = rules["func"]
    enable_value = rules["enable"]
    int_offset = rules["int_offset"]
    int_start = rules["int_start"]

    rows = []
    global_block_idx = 0
    prev_intaddr = int_start
    prev_count = None
    prev_device = None

    for device_no in range(1, devices + 1):
        node_no = node_seq[device_no - 1]
        for block_no in range(1, blocks + 1):
            idx = global_block_idx
            for p in param_template:
                param = p.format(i=idx)
                base = param.split('.')[-1]
                cfg = ""

                if block_no in count_map or block_no in devaddr_map:
                    if base == "Enable":
                        cfg = enable_value
                    elif base == "Count":
                        cfg = count_map.get(block_no, "")
                    elif base == "IntAddress":
                        if prev_intaddr is None:
                            cfg = int_start
                        else:
                            addition = prev_intaddr + (prev_count if prev_count else 0)
                            if prev_device and prev_device != device_no:
                                addition += int_offset
                            cfg = addition
                    elif base == "Node":
                        cfg = node_no
                    elif base == "Func":
                        cfg = func_value
                    elif base == "DevAddress":
                        cfg = devaddr_map.get(block_no, "")

                rows.append({
                    "Device No.": device_no,
                    "Block No.": block_no,
                    "Node No.": node_no,
                    "Parameter": param,
                    "ConfigValue": cfg
                })

            # update counters
            if block_no in count_map:
                prev_count = count_map[block_no]
                last_8 = rows[-8:]
                assigned_intaddr = last_8[1]['ConfigValue']
                prev_intaddr = assigned_intaddr
                prev_device = device_no

            global_block_idx += 1

    df = pd.DataFrame(rows)

    # Save to memory buffer
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Port1", index=False)
        pd.DataFrame(columns=df.columns).to_excel(writer, sheet_name="Port2", index=False)

    buffer.seek(0)

    # === Apply Formatting (colors + thick borders) ===
    wb = load_workbook(buffer)
    ws = wb["Port1"]

    thin = Side(border_style="thick", color="000000")

    # alternating block colors
    block_colors = ["FFF2CC", "D9EAD3", "FCE5CD", "EAD1DC", "D0E0E3", "C9DAF8"]

    start_row = 2
    row_idx = start_row
    for device_no in range(1, devices + 1):
        for block_no in range(1, blocks + 1):
            block_start = row_idx
            block_end = block_start + rows_per_block - 1
            fill_color = block_colors[(block_no - 1) % len(block_colors)]
            fill = PatternFill(fill_type="solid", fgColor=fill_color)

            # Apply fill and border
            for r in range(block_start, block_end + 1):
                for c in range(1, 6):
                    ws.cell(row=r, column=c).fill = fill
                    # Apply thick borders around block
                    if r == block_start:
                        ws.cell(row=r, column=c).border = Border(top=thin)
                    if r == block_end:
                        ws.cell(row=r, column=c).border = Border(bottom=thin)
                    if c == 1:
                        ws.cell(row=r, column=c).border = Border(left=thin)
                    if c == 5:
                        ws.cell(row=r, column=c).border = Border(right=thin)

            row_idx += rows_per_block

    final_buf = BytesIO()
    wb.save(final_buf)
    final_buf.seek(0)
    return final_buf


# === Streamlit Web App ===
st.title("⚙️ MasterCmd Excel Generator (Web Version)")

devices = st.number_input("Number of Devices", min_value=1, value=26)
blocks = st.number_input("Blocks per Device", min_value=1, value=6)
rows_per_block = st.number_input("Rows per Block", min_value=1, value=8)
node_str = st.text_input("Node Numbers (comma-separated)", "26,27,28,29,30,31,32,33,34,35,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76")

st.subheader("Rule Configuration")
enable_value = st.number_input(".Enable Value", value=1)
func_value = st.number_input(".Func Value", value=3)
int_offset = st.number_input("IntAddress Offset (new device)", value=10)
int_start = st.number_input("Initial .IntAddress Value", value=0)

st.markdown("#### Block Mappings")
count_map = {}
devaddr_map = {}

for b in range(1, blocks + 1):
    col1, col2, col3 = st.columns([1, 1, 1])
    with col1:
        st.text(f"Block {b}")
    with col2:
        cval = st.text_input(f".Count Value {b}", value=str(1 if b in [1, 3, 4] else 5), key=f"count{b}")
        if cval:
            count_map[b] = int(cval)
    with col3:
        dval = st.text_input(f".DevAddress Value {b}", value=str({1:3, 2:101, 3:116, 4:142}.get(b, "")), key=f"devaddr{b}")
        if dval:
            devaddr_map[b] = int(dval)

node_seq = [int(x.strip()) for x in node_str.split(",")]
rules = {
    "count_map": count_map,
    "devaddr_map": devaddr_map,
    "enable": enable_value,
    "func": func_value,
    "int_offset": int_offset,
    "int_start": int_start
}

if st.button("Generate Excel File"):
    if len(node_seq) != devices:
        st.error("❌ Number of node numbers must match number of devices.")
    else:
        buffer = generate_excel(devices, blocks, rows_per_block, node_seq, rules)
        st.success("✅ Excel file generated successfully!")
        st.download_button(
            label="📥 Download Excel",
            data=buffer,
            file_name="MasterCmd_Sequence.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
